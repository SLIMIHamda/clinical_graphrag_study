import openpyxl
import pytest

from mgr.eval.b0_crosscheck import crosscheck, load_reference
from mgr.figures import plot_coverage_curve, plot_pareto, plot_rgd
from mgr.rerank.care_gate import FrontierPoint
from mgr.stats.rgd import decompose
from mgr.sync_status import statuses_from_results, write_statuses_to_xlsx
from mgr.tracking import record as rec


# ----- sync_status --------------------------------------------------------- #

def test_statuses_from_results(tmp_path):
    root = tmp_path / "results"
    r = rec.RunRecord(
        run_id="R0001", slug="s", condition="No-RAG", benchmark="MMLU-Med",
        backbone="Llama-70B", seed=42, status="Done", config_hash="h",
    )
    rec.log(r, root)
    assert statuses_from_results(root) == {"R0001": "Done"}


def test_write_statuses_to_xlsx_roundtrip(tmp_path):
    xlsx = tmp_path / "m.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Run_Manifest"
    ws.append(["run_id", "status"])
    ws.append(["R0001", "Pending"])
    ws.append(["R0002", "Pending"])
    wb.save(xlsx)

    changed = write_statuses_to_xlsx(xlsx, {"R0001": "Done"})
    assert changed == 1
    wb2 = openpyxl.load_workbook(xlsx)
    rows = list(wb2["Run_Manifest"].iter_rows(values_only=True))
    assert ("R0001", "Done") in rows
    assert ("R0002", "Pending") in rows


# ----- B0 cross-check ------------------------------------------------------ #

def test_b0_crosscheck_pass_and_fail():
    ref = {"MMLU-Med|No-RAG": 0.78, "MedQA-US|No-RAG": 0.70}
    obs_ok = {"MMLU-Med|No-RAG": 0.80, "MedQA-US|No-RAG": 0.68}
    assert crosscheck(obs_ok, ref, tolerance=0.05).passed
    obs_bad = {"MMLU-Med|No-RAG": 0.60, "MedQA-US|No-RAG": 0.70}
    rep = crosscheck(obs_bad, ref, tolerance=0.05)
    assert not rep.passed
    assert "FAIL" in rep.summary()


def test_b0_load_reference_skips_notes(tmp_path):
    p = tmp_path / "ref.json"
    p.write_text('{"_note": "hi", "MMLU-Med|No-RAG": 0.78}', encoding="utf-8")
    ref = load_reference(p)
    assert ref == {"MMLU-Med|No-RAG": 0.78}


# ----- figures (smoke: a non-empty file is produced) ----------------------- #

def test_figures_write_files(tmp_path):
    pts = decompose({"BM25": (0.5, 0.6), "Graph-only": (0.7, 0.55)}, baseline="BM25")
    f1 = plot_rgd(pts, tmp_path / "f3.png", baseline="BM25")
    assert f1.exists() and f1.stat().st_size > 0

    frontier = {
        "care": FrontierPoint("care", 0.5, 0.8, 2.0),
        "always": FrontierPoint("always", 1.0, 0.8, 4.0),
        "never": FrontierPoint("never", 0.0, 0.6, 0.0),
    }
    f2 = plot_pareto(frontier, tmp_path / "f4.png")
    assert f2.exists() and f2.stat().st_size > 0

    f3 = plot_coverage_curve({"exact": 0.5, "exact+abbrev": 0.65, "exact+abbrev+fuzzy": 0.8}, tmp_path / "f5.png")
    assert f3.exists() and f3.stat().st_size > 0
