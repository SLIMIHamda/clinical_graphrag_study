"""sync_status — reconcile run-records back into the workbook status column.

Run-records are the source of truth for *what happened* (Doc 00 section 0). This
reads the per-run records, derives each run's status, and writes it back into the
xlsx Run_Manifest ``status`` column so Budget_Summary's live progress reflects
reality. The xlsx is opened normally (not data_only) to preserve formulas.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl


def statuses_from_results(results_root: str | Path) -> dict[str, str]:
    """Map run_id -> status from the per-run JSON records."""
    per_run = Path(results_root) / "per-run"
    out: dict[str, str] = {}
    if not per_run.exists():
        return out
    for jf in sorted(per_run.glob("R*.json")):
        try:
            d = json.loads(jf.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        if "run_id" in d and "status" in d:
            out[str(d["run_id"])] = str(d["status"])
    return out


def write_statuses_to_xlsx(
    xlsx_path: str | Path,
    statuses: dict[str, str],
    *,
    sheet: str = "Run_Manifest",
    id_col: str = "run_id",
    status_col: str = "status",
) -> int:
    """Update the status column for matching run_ids. Returns rows changed."""
    wb = openpyxl.load_workbook(xlsx_path)  # keep formulas
    ws = wb[sheet]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    try:
        id_idx = header.index(id_col)
        st_idx = header.index(status_col)
    except ValueError as e:
        raise KeyError(f"column not found in {sheet}: {e}")

    changed = 0
    for row in ws.iter_rows(min_row=2):
        rid = row[id_idx].value
        if rid is None:
            continue
        rid = str(rid)
        if rid in statuses and row[st_idx].value != statuses[rid]:
            row[st_idx].value = statuses[rid]
            changed += 1
    wb.save(xlsx_path)
    return changed


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Reconcile run-records -> xlsx status column")
    ap.add_argument("--xlsx", default="manifest/Experiment-Matrix.xlsx")
    ap.add_argument("--results-root", default="results")
    args = ap.parse_args(argv)

    statuses = statuses_from_results(args.results_root)
    if not statuses:
        print("no run-records found", file=sys.stderr)
        return 1
    changed = write_statuses_to_xlsx(args.xlsx, statuses)
    done = sum(1 for s in statuses.values() if s == "Done")
    print(f"synced {len(statuses)} records ({done} Done); updated {changed} rows in {args.xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
